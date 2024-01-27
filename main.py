import argparse
import json
from tools.template_parser import Template
from openpyxl import load_workbook
from tools.teacher_parser import TeacherInfo
from tools.image_maker import image_maker, make_dark_image
import os
from tqdm import tqdm

if __name__ == '__main__':
    parser = argparse.ArgumentParser(
        prog='TimeTableImager',
        description='Create timetable image from xlsx file.',
        epilog='Made by @seolmango'
    )

    parser.add_argument("-s", "--setting", type=str, default="setting.json", help="setting file path")
    parser.add_argument("-i", "--input", type=str, default="table.xlsx", help="input xlsx file path")
    parser.add_argument("-sh", "--sheet", type=str, default="Sheet1", help="sheet name of xlsx file")
    parser.add_argument("-o", "--output", type=str, default="output", help="output directory path")
    parser.add_argument("-n", "--name", type=str, default="<auto>", help="output image name")
    parser.add_argument("-f", "--format", type=str, default="png", help="output image format")
    parser.add_argument("-sl", "--start_line", type=int, default=0, help="start line of timetable in xlsx file")
    parser.add_argument("-t", "--template", type=str, default="template.txt", help="template file path")
    parser.add_argument('-e', '--extra-data', type=str, default='', help='extra data file path')
    args = parser.parse_args()

    with open(args.setting, 'r', encoding='utf-8') as f:
        setting = json.load(f)

    template = Template(setting['Template'] if 'Template' in setting else args.template)

    wb = load_workbook(setting['TableData'] if 'TableData' in setting else args.input)
    ws = wb[setting['Sheet'] if 'Sheet' in setting else args.sheet]

    teacher_info = TeacherInfo(setting['ExtraData'] if 'ExtraData' in setting else args.extra_data)

    table_data = []
    count = 0
    for i in range(setting['StartLine'] if 'StartLine' in setting else args.start_line, ws.max_row, template.temp_height):
        ws_crop = []
        for j in range(i, i + template.temp_height):
            ws_crop.append([])
            for k in range(template.temp_width):
                ws_crop[-1].append(ws.cell(row=j, column=k+1))
        temp = template.parser(ws_crop)
        for j in temp.keys():
            if j != 'name' and temp[j]['name'] != None:
                temp[j]['teacher'] = teacher_info.get_info(j, temp[j]['name'][:-3], temp[j]['class'])['name']
                temp[j]['location'] = teacher_info.get_info(j, temp[j]['name'][:-3], temp[j]['class'])['location']

        table_data.append(temp)
        table_data[-1]['count'] = count
        count += 1

    if not os.path.exists(setting['Output'] if 'Output' in setting else args.output):
        os.mkdir(setting['Output'] if 'Output' in setting else args.output)

    for i in tqdm(table_data):
        name = i['name']['name']
        image_maker(i, f'{setting["Output"] if "Output" in setting else args.output}/{name if name != None else args.name}.{setting["Format"] if "Format" in setting else args.format}')
        make_dark_image(i, f'{setting["Output"] if "Output" in setting else args.output}/{name if name != None else args.name}_dark.{setting["Format"] if "Format" in setting else args.format}')