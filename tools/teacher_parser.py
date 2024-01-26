from openpyxl import load_workbook

class TeacherInfo:
    def __init__(self, file_path):
        row_data = load_workbook(file_path)
        row_data_sh = row_data['2학기 강의실']
        self.one_teacher_class_data = {}
        teacher_class_loc = {}
        for i in range(2, row_data_sh.max_row):
            if row_data_sh.cell(row=i, column=5).value == None:
                self.one_teacher_class_data[row_data_sh.cell(row=i, column=2).value] = {
                    "name": row_data_sh.cell(row=i, column=3).value,
                    "location": row_data_sh.cell(row=i, column=4).value
                }
            else:
                for j in [3,5,7]:
                    if row_data_sh.cell(row=i, column=j).value != None and row_data_sh.cell(row=i, column=j).value.strip() != '':
                        teacher_class_loc[f'{row_data_sh.cell(row=i, column=j).value} {row_data_sh.cell(row=i, column=2).value}'] = row_data_sh.cell(row=i, column=j+1).value
        row_data_sh_2 = row_data['과목별 다교사수업']
        self.two_teacher_class_data = {}
        temp_sd = {4: 'mon', 5: 'tue', 6: 'wed', 7: 'thu', 8: 'fri'}
        for i in range(2, row_data_sh_2.max_row):
            teacher_name = row_data_sh_2.cell(row=i, column=3).value
            class_name = row_data_sh_2.cell(row=i, column=2).value
            for j in [4,5,6,7,8]:
                if row_data_sh_2.cell(row=i, column=j).value != None:
                    temp = row_data_sh_2.cell(row=i, column=j).value.split('/')
                    for k in temp:
                        sd = k.split('(')
                        times = sd[0].split(',')
                        for time in times:
                            if f'{temp_sd[j]}_{time}' in self.two_teacher_class_data:
                                self.two_teacher_class_data[f'{temp_sd[j]}_{time}'].append({
                                    "name": teacher_name,
                                    "class_name": f'{class_name}',
                                    "class_num": f"{sd[1][0]}반",
                                    "location": teacher_class_loc[f'{teacher_name} {class_name}']
                                })
                            else:
                                self.two_teacher_class_data[f'{temp_sd[j]}_{time}'] = [{
                                    "name": teacher_name,
                                    "class_name": f'{class_name}',
                                    "class_num": f"{sd[1][0]}반",
                                    "location": teacher_class_loc[f'{teacher_name} {class_name}']
                                }]

    def get_info(self, time, class_name, class_num):
        for i in self.one_teacher_class_data:
            if class_name == i:
                return self.one_teacher_class_data[i]
        for i in self.two_teacher_class_data[time]:
            if class_name.startswith(i['class_name']) and class_num == i['class_num']:
                return i
        return {
            "name": None,
            "location": None
        }

